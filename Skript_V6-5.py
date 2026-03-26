from __future__ import annotations
import uuid
from pathlib import Path
from typing import Dict, List, Tuple, Optional, TypedDict, Any

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "vorlage.xlsx"
SKIP_SHEETS = {"Befehle"}

SHEETS: Optional[List[str]] = None

_TARGET_W = 3860
_TARGET_H = 2400
_BASE_W = 2000
_BASE_H = 1540
_SCALE_X = _TARGET_W / _BASE_W
_SCALE_Y = _TARGET_H / _BASE_H
_OFFSET_X = 0

ROOT_FRAME = (0, 0, _TARGET_W, _TARGET_H)
TABBAR_SIZE = int(round(40 * _SCALE_Y))

MARGIN_X = _OFFSET_X + int(round(5 * _SCALE_X))
MARGIN_Y = int(round(5 * _SCALE_Y))
GAP_X = int(round(5 * _SCALE_X))
GAP_Y = int(round(5 * _SCALE_Y))
COLUMNS = 10
BUTTON_W = int(round(190 * _SCALE_X))
BUTTON_H = int(round(140 * _SCALE_Y))

TEXT_HEIGHT = int(round(120 * _SCALE_Y))
TEXT_PADDING_X = int(round(10 * _SCALE_X))
TEXT_PADDING_Y = int(round(10 * _SCALE_Y))
DEFAULT_TEXT_SIZE = 40
HEADER_OSC_PATH = "OSC FERTIG"
HEADER_LABEL = "BUTTON LABEL"

HEADER_TEXT_SIZE = "Schriftgröße"
HEADER_LABEL_COLOR = "FARBE LABEL"
HEADER_BUTTON_COLOR = "FARBE BUTTON"

HEADER_MIDI_CHANNEL = "MIDI CHANNEL"
HEADER_MIDI_CONTROLLER = "MIDI Controller"
HEADER_MIDI_ENABLE = "MIDI ENABLE"
HEADER_MIDI_SEND = "MIDI SEND"
HEADER_MIDI_RECEIVE = "MIDI RECEIVE"

HEADER_OSC_ENABLE = "OSC ENABLE"
HEADER_OSC_SEND = "OSC SEND"
HEADER_OSC_RECEIVE = "OSC RECEIVE"

HEADER_TAB_COLOR_CANDIDATES = ("Reiter Farbe", "Farbe")

DEFAULT_LABEL_COLOR = "FFFFFFFF"
DEFAULT_BUTTON_COLOR = "D3D3D3FF"
DEFAULT_TAB_COLOR = "FF0000FF"

USE_LABEL_AS_NODE_NAME = True

OUT_FILENAME = "touchosc_from_xlsx.tosc"

class RowDef(TypedDict):
    position: int
    osc_path: str
    label: str
    label_size: int
    label_color: str
    button_color: str
    midi_channel: int
    midi_controller: int
    midi_enable: int
    midi_send: int
    midi_receive: int
    osc_enable: int
    osc_send: int
    osc_receive: int

def new_id() -> str:
    return str(uuid.uuid4())

def find_header_indices(header_row: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, v in enumerate(header_row):
        if v is None:
            continue
        txt = str(v).strip()
        if txt:
            idx[txt] = i
    return idx

def _to_int(value: Any, default: int) -> int:
    try:
        if value is None:
            return default
        s = str(value).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default

def _yn_to_int(value: Any) -> int:
    if value is None:
        return 0
    s = str(value).strip().lower()
    return 1 if s == "y" else 0

def _normalize_hex_color(value: Any, default: str) -> str:
    """Normalisiert Farbcodes zu RRGGBBAA (8 Hex-Zeichen).
    Akzeptiert:
      - 'FF0000FF' / '#FF0000FF' (RRGGBBAA)
      - 'FF0000'   / '#FF0000'   (RRGGBB -> Alpha = FF)
    """
    if value is None:
        return default

    s = str(value).strip()
    if not s:
        return default

    if s.startswith("#"):
        s = s[1:].strip()

    s = s.upper()

    import re
    if not re.fullmatch(r"[0-9A-F]{6}([0-9A-F]{2})?", s):
        return default

    if len(s) == 6:
        return s + "FF"
    return s

def _hex_to_rgba_floats(rrggbbaa: str, default: str) -> Tuple[float, float, float, float]:
    rrggbbaa = _normalize_hex_color(rrggbbaa, default)
    r = int(rrggbbaa[0:2], 16) / 255.0
    g = int(rrggbbaa[2:4], 16) / 255.0
    b = int(rrggbbaa[4:6], 16) / 255.0
    a = int(rrggbbaa[6:8], 16) / 255.0
    return r, g, b, a

def _osc_message_xml(*, osc_path: str, enabled: int, send: int, receive: int) -> str:
    return f"""                                    <osc>
                                        <enabled>{enabled}</enabled>
                                        <send>{send}</send>
                                        <receive>{receive}</receive>
                                        <feedback>0</feedback>
                                        <noDuplicates>0</noDuplicates>
                                        <connections>1111111111</connections>
                                        <triggers>
                                            <trigger><var><![CDATA[x]]></var><condition>ANY</condition></trigger>
                                        </triggers>
                                        <path>
                                            <partial>
                                                <type>CONSTANT</type>
                                                <conversion>STRING</conversion>
                                                <value><![CDATA[{osc_path}]]></value>
                                                <scaleMin>0</scaleMin>
                                                <scaleMax>1</scaleMax>
                                            </partial>
                                        </path>
                                        <arguments></arguments>
                                    </osc>
"""

def _midi_message_xml(*, channel: int, controller: int, enabled: int, send: int, receive: int) -> str:
    """Erzeugt einen TouchOSC-MIDI-Block für Control Change.

    TouchOSC erwartet hier die Struktur:
      <midi> ... <message><type>CONTROLCHANGE</type><channel>..</channel><data1>..</data1><data2>..</data2></message>
           <values> (channel, data1, data2) </values>
      </midi>

    - channel: aus Excel 1..16 -> TouchOSC intern 0..15
    - controller: 0..127 (CC-Nummer) -> data1
    - value: VALUE aus Button-Value 'x' -> data2, skaliert 0..127
    """
    ch0 = max(0, min(15, int(channel) - 1))
    cc = max(0, min(127, int(controller)))

    return f"""                                    <midi>
                                        <enabled>{enabled}</enabled>
                                        <send>{send}</send>
                                        <receive>{receive}</receive>
                                        <feedback>0</feedback>
                                        <noDuplicates>0</noDuplicates>
                                        <connections>1111111111</connections>
                                        <triggers>
                                            <trigger><var><![CDATA[x]]></var><condition>ANY</condition></trigger>
                                        </triggers>
                                        <message>
                                            <type>CONTROLCHANGE</type>
                                            <channel>{ch0}</channel>
                                            <data1>{cc}</data1>
                                            <data2>0</data2>
                                        </message>
                                        <values>
                                            <value>
                                                <type>CONSTANT</type>
                                                <key><![CDATA[]]></key>
                                                <scaleMin>0</scaleMin>
                                                <scaleMax>15</scaleMax>
                                            </value>
                                            <value>
                                                <type>CONSTANT</type>
                                                <key><![CDATA[]]></key>
                                                <scaleMin>0</scaleMin>
                                                <scaleMax>127</scaleMax>
                                            </value>
                                            <value>
                                                <type>VALUE</type>
                                                <key><![CDATA[x]]></key>
                                                <scaleMin>0</scaleMin>
                                                <scaleMax>127</scaleMax>
                                            </value>
                                        </values>
                                    </midi>
"""

def read_rows_from_ws(ws) -> Tuple[List[RowDef], str]:
    """Liest pro Zeile: RowDef und zusätzlich eine Tab-Farbe (optional aus dem Sheet)."""
    rows = list(ws.iter_rows(values_only=True))

    header_row_index = None
    header_map: Optional[Dict[str, int]] = None

    for r_i, row in enumerate(rows):
        header_candidates = [str(c).strip() if c is not None else "" for c in row]
        m = find_header_indices(header_candidates)
        if HEADER_OSC_PATH in m and HEADER_LABEL in m:
            header_row_index = r_i
            header_map = m
            break

    if header_row_index is None or header_map is None:
        raise RuntimeError(
            f"[{ws.title}] Konnte Header nicht finden. Erwartet Spalten '{HEADER_OSC_PATH}' und '{HEADER_LABEL}'."
        )

    def idx(name: str) -> Optional[int]:
        return header_map.get(name)

    i_osc = idx(HEADER_OSC_PATH)
    i_label = idx(HEADER_LABEL)

    i_size = idx(HEADER_TEXT_SIZE)
    i_label_color = idx(HEADER_LABEL_COLOR)
    i_button_color = idx(HEADER_BUTTON_COLOR)

    i_midi_ch = idx(HEADER_MIDI_CHANNEL)
    i_midi_cc = idx(HEADER_MIDI_CONTROLLER)
    i_midi_en = idx(HEADER_MIDI_ENABLE)
    i_midi_send = idx(HEADER_MIDI_SEND)
    i_midi_recv = idx(HEADER_MIDI_RECEIVE)

    i_osc_en = idx(HEADER_OSC_ENABLE)
    i_osc_send = idx(HEADER_OSC_SEND)
    i_osc_recv = idx(HEADER_OSC_RECEIVE)

    tab_color = DEFAULT_TAB_COLOR
    tab_col_index: Optional[int] = None
    for cand in HEADER_TAB_COLOR_CANDIDATES:
        if cand in header_map:
            tab_col_index = header_map[cand]
            break
    if tab_col_index is not None:
        for row in rows[header_row_index + 1:]:
            v = row[tab_col_index] if tab_col_index < len(row) else None
            norm = _normalize_hex_color(v, "")
            if norm:
                tab_color = norm
                break

    out: List[RowDef] = []
    assert i_osc is not None and i_label is not None

    for pos, row in enumerate(rows[header_row_index + 1:], start=1):
        label = row[i_label] if i_label < len(row) else None

        label_s = str(label).strip() if label is not None else ""
        if not label_s:
            continue

        osc = row[i_osc] if i_osc < len(row) else None
        osc_s = str(osc).strip() if osc is not None else ""

        label_size = _to_int(row[i_size] if (i_size is not None and i_size < len(row)) else None, DEFAULT_TEXT_SIZE)

        label_color = _normalize_hex_color(
            row[i_label_color] if (i_label_color is not None and i_label_color < len(row)) else None,
            DEFAULT_LABEL_COLOR,
        )

        button_color = _normalize_hex_color(
            row[i_button_color] if (i_button_color is not None and i_button_color < len(row)) else None,
            DEFAULT_BUTTON_COLOR,
        )

        midi_channel = _to_int(row[i_midi_ch] if (i_midi_ch is not None and i_midi_ch < len(row)) else None, 1)
        if midi_channel < 1:
            midi_channel = 1
        if midi_channel > 16:
            midi_channel = 16

        midi_controller = _to_int(row[i_midi_cc] if (i_midi_cc is not None and i_midi_cc < len(row)) else None, 0)
        if midi_controller < 0:
            midi_controller = 0
        if midi_controller > 127:
            midi_controller = 127

        midi_enable = _yn_to_int(row[i_midi_en] if (i_midi_en is not None and i_midi_en < len(row)) else None)
        midi_send = _yn_to_int(row[i_midi_send] if (i_midi_send is not None and i_midi_send < len(row)) else None)
        midi_receive = _yn_to_int(row[i_midi_recv] if (i_midi_recv is not None and i_midi_recv < len(row)) else None)

        osc_enable = _yn_to_int(row[i_osc_en] if (i_osc_en is not None and i_osc_en < len(row)) else None)
        osc_send = _yn_to_int(row[i_osc_send] if (i_osc_send is not None and i_osc_send < len(row)) else None)
        osc_receive = _yn_to_int(row[i_osc_recv] if (i_osc_recv is not None and i_osc_recv < len(row)) else None)

        out.append(
            RowDef(
                position=pos,
                osc_path=osc_s,
                label=label_s,
                label_size=label_size,
                label_color=label_color,
                button_color=button_color,
                midi_channel=midi_channel,
                midi_controller=midi_controller,
                midi_enable=midi_enable,
                midi_send=midi_send,
                midi_receive=midi_receive,
                osc_enable=osc_enable,
                osc_send=osc_send,
                osc_receive=osc_receive,
            )
        )

    if not out:
        raise RuntimeError(f"[{ws.title}] Keine gültigen OSC-Pfade gefunden (Spalte '{HEADER_OSC_PATH}').")

    return out, tab_color

def make_multiline_text(label: str) -> str:
    parts = label.strip().split()
    if len(parts) >= 3 and parts[0].upper() == "GROUP" and parts[1].isdigit() and parts[2].isdigit():
        return f"GROUP\n {parts[1]} / {parts[2]}\n"
    return label.strip()

def text_node_xml(*, node_id: str, name: str, x: int, y: int, w: int, h: int, text: str, text_size: int, text_color_hex: str) -> str:
    r, g, b, a = _hex_to_rgba_floats(text_color_hex, DEFAULT_LABEL_COLOR)
    return f"""                            <node ID='{node_id}' type='TEXT'>
                                <properties>
                                    <property type='b'><key><![CDATA[background]]></key><value>0</value></property>
                                    <property type='c'><key><![CDATA[color]]></key><value><r>0.6</r><g>0.6</g><b>0.6</b><a>1</a></value></property>
                                    <property type='f'><key><![CDATA[cornerRadius]]></key><value>1</value></property>
                                    <property type='i'><key><![CDATA[font]]></key><value>0</value></property>
                                    <property type='r'><key><![CDATA[frame]]></key><value><x>{x}</x><y>{y}</y><w>{w}</w><h>{h}</h></value></property>
                                    <property type='b'><key><![CDATA[grabFocus]]></key><value>0</value></property>
                                    <property type='b'><key><![CDATA[interactive]]></key><value>0</value></property>
                                    <property type='b'><key><![CDATA[locked]]></key><value>0</value></property>
                                    <property type='s'><key><![CDATA[name]]></key><value><![CDATA[{name}]]></value></property>
                                    <property type='i'><key><![CDATA[orientation]]></key><value>0</value></property>
                                    <property type='b'><key><![CDATA[outline]]></key><value>0</value></property>
                                    <property type='i'><key><![CDATA[outlineStyle]]></key><value>0</value></property>
                                    <property type='i'><key><![CDATA[pointerPriority]]></key><value>0</value></property>
                                    <property type='i'><key><![CDATA[shape]]></key><value>1</value></property>
                                    <property type='i'><key><![CDATA[textAlignH]]></key><value>2</value></property>
                                    <property type='i'><key><![CDATA[textAlignV]]></key><value>2</value></property>
                                    <property type='b'><key><![CDATA[textClip]]></key><value>1</value></property>
                                    <property type='c'><key><![CDATA[textColor]]></key><value><r>{r}</r><g>{g}</g><b>{b}</b><a>{a}</a></value></property>
                                    <property type='i'><key><![CDATA[textSize]]></key><value>{text_size}</value></property>
                                    <property type='b'><key><![CDATA[textWrap]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[visible]]></key><value>1</value></property>
                                </properties>
                                <values>
                                    <value>
                                        <key><![CDATA[text]]></key>
                                        <locked>0</locked>
                                        <lockedDefaultCurrent>1</lockedDefaultCurrent>
                                        <default><![CDATA[{text}]]></default>
                                        <defaultPull>0</defaultPull>
                                    </value>
                                    <value>
                                        <key><![CDATA[touch]]></key>
                                        <locked>0</locked>
                                        <lockedDefaultCurrent>0</lockedDefaultCurrent>
                                        <default><![CDATA[false]]></default>
                                        <defaultPull>0</defaultPull>
                                    </value>
                                </values>
                            </node>
"""

def button_node_xml(
    *,
    node_id: str,
    name: str,
    x: int,
    y: int,
    w: int,
    h: int,
    osc_path: str,
    button_color_hex: str,
    midi_channel: int,
    midi_controller: int,
    midi_enable: int,
    midi_send: int,
    midi_receive: int,
    osc_enable: int,
    osc_send: int,
    osc_receive: int,
) -> str:
    r, g, b, a = _hex_to_rgba_floats(button_color_hex, DEFAULT_BUTTON_COLOR)

    messages = ""
    messages += _osc_message_xml(osc_path=osc_path, enabled=osc_enable, send=osc_send, receive=osc_receive)
    messages += _midi_message_xml(channel=midi_channel, controller=midi_controller, enabled=midi_enable, send=midi_send, receive=midi_receive)

    return f"""                            <node ID='{node_id}' type='BUTTON'>
                                <properties>
                                    <property type='b'><key><![CDATA[background]]></key><value>1</value></property>
                                    <property type='i'><key><![CDATA[buttonType]]></key><value>0</value></property>
                                    <property type='c'><key><![CDATA[color]]></key><value><r>{r}</r><g>{g}</g><b>{b}</b><a>{a}</a></value></property>
                                    <property type='f'><key><![CDATA[cornerRadius]]></key><value>1</value></property>
                                    <property type='r'><key><![CDATA[frame]]></key><value><x>{x}</x><y>{y}</y><w>{w}</w><h>{h}</h></value></property>
                                    <property type='b'><key><![CDATA[grabFocus]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[interactive]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[locked]]></key><value>0</value></property>
                                    <property type='s'><key><![CDATA[name]]></key><value><![CDATA[{name}]]></value></property>
                                    <property type='i'><key><![CDATA[orientation]]></key><value>0</value></property>
                                    <property type='b'><key><![CDATA[outline]]></key><value>0</value></property>
                                    <property type='i'><key><![CDATA[outlineStyle]]></key><value>1</value></property>
                                    <property type='i'><key><![CDATA[pointerPriority]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[press]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[release]]></key><value>1</value></property>
                                    <property type='i'><key><![CDATA[shape]]></key><value>1</value></property>
                                    <property type='b'><key><![CDATA[valuePosition]]></key><value>0</value></property>
                                    <property type='b'><key><![CDATA[visible]]></key><value>1</value></property>
                                </properties>
                                <values>
                                    <value><key><![CDATA[x]]></key><locked>0</locked><lockedDefaultCurrent>0</lockedDefaultCurrent><default><![CDATA[0]]></default><defaultPull>0</defaultPull></value>
                                    <value><key><![CDATA[touch]]></key><locked>0</locked><lockedDefaultCurrent>0</lockedDefaultCurrent><default><![CDATA[false]]></default><defaultPull>0</defaultPull></value>
                                </values>
                                <messages>
{messages}                                </messages>
                            </node>
"""

def build_page_children_xml(defs: List[RowDef]) -> str:
    children_xml: List[str] = []

    for d in defs:
        idx = d["position"]
        col = (idx - 1) % COLUMNS
        row = (idx - 1) // COLUMNS

        bx = MARGIN_X + col * (BUTTON_W + GAP_X)
        by = MARGIN_Y + row * (BUTTON_H + GAP_Y)

        node_name = d["label"] if USE_LABEL_AS_NODE_NAME else f"button{idx}"

        tx = bx + TEXT_PADDING_X
        ty = by + TEXT_PADDING_Y
        tw = BUTTON_W - 2 * TEXT_PADDING_X
        th = TEXT_HEIGHT

        children_xml.append(
            button_node_xml(
                node_id=new_id(),
                name=node_name,
                x=bx,
                y=by,
                w=BUTTON_W,
                h=BUTTON_H,
                osc_path=d["osc_path"],
                button_color_hex=d["button_color"],
                midi_channel=d["midi_channel"],
                midi_controller=d["midi_controller"],
                midi_enable=d["midi_enable"],
                midi_send=d["midi_send"],
                midi_receive=d["midi_receive"],
                osc_enable=d["osc_enable"],
                osc_send=d["osc_send"],
                osc_receive=d["osc_receive"],
            )
        )
        children_xml.append(
            text_node_xml(
                node_id=new_id(),
                name=f"text{idx}",
                x=tx,
                y=ty,
                w=tw,
                h=th,
                text=make_multiline_text(d["label"]),
                text_size=d["label_size"],
                text_color_hex=d["label_color"],
            )
        )

    return "".join(children_xml)

def page_group_xml(*, page_id: str, page_name: str, tab_label: str, tab_color_hex: str, inner_children: str) -> str:
    rx, ry, rw, rh = ROOT_FRAME
    page_frame_y = TABBAR_SIZE
    page_frame_h = rh - TABBAR_SIZE

    tr, tg, tb, ta = _hex_to_rgba_floats(tab_color_hex, DEFAULT_TAB_COLOR)

    return f"""                    <node ID='{page_id}' type='GROUP'>
                        <properties>
                            <property type='b'><key><![CDATA[background]]></key><value>1</value></property>
                            <property type='c'><key><![CDATA[color]]></key><value><r>0</r><g>0</g><b>0</b><a>0</a></value></property>
                            <property type='f'><key><![CDATA[cornerRadius]]></key><value>1</value></property>
                            <property type='r'><key><![CDATA[frame]]></key><value><x>0</x><y>{page_frame_y}</y><w>{rw}</w><h>{page_frame_h}</h></value></property>
                            <property type='b'><key><![CDATA[grabFocus]]></key><value>0</value></property>
                            <property type='b'><key><![CDATA[interactive]]></key><value>0</value></property>
                            <property type='b'><key><![CDATA[locked]]></key><value>0</value></property>
                            <property type='s'><key><![CDATA[name]]></key><value><![CDATA[{page_name}]]></value></property>
                            <property type='i'><key><![CDATA[orientation]]></key><value>0</value></property>
                            <property type='b'><key><![CDATA[outline]]></key><value>0</value></property>
                            <property type='i'><key><![CDATA[outlineStyle]]></key><value>0</value></property>
                            <property type='i'><key><![CDATA[pointerPriority]]></key><value>0</value></property>
                            <property type='i'><key><![CDATA[shape]]></key><value>1</value></property>

                            <property type='s'><key><![CDATA[tabLabel]]></key><value><![CDATA[{tab_label}]]></value></property>
                            <property type='c'><key><![CDATA[tabColorOff]]></key><value><r>{tr}</r><g>{tg}</g><b>{tb}</b><a>{ta}</a></value></property>
                            <property type='c'><key><![CDATA[tabColorOn]]></key><value><r>0.0</r><g>0.2</g><b>0.6</b><a>1</a></value></property>
                            <property type='c'><key><![CDATA[textColorOff]]></key><value><r>1</r><g>1</g><b>1</b><a>1</a></value></property>
                            <property type='c'><key><![CDATA[textColorOn]]></key><value><r>1</r><g>1</g><b>1</b><a>1</a></value></property>

                            <property type='b'><key><![CDATA[visible]]></key><value>1</value></property>
                        </properties>

                        <values>
                            <value>
                                <key><![CDATA[touch]]></key>
                                <locked>0</locked>
                                <lockedDefaultCurrent>0</lockedDefaultCurrent>
                                <default><![CDATA[false]]></default>
                                <defaultPull>0</defaultPull>
                            </value>
                        </values>

                        <children>
{inner_children}                        </children>
                    </node>
"""

def pager_xml(*, pager_id: str, pages_xml: str) -> str:
    rx, ry, rw, rh = ROOT_FRAME
    return f"""            <node ID='{pager_id}' type='PAGER'>
                <properties>
                    <property type='b'><key><![CDATA[background]]></key><value>1</value></property>
                    <property type='c'><key><![CDATA[color]]></key><value><r>0.25</r><g>0.25</g><b>0.25</b><a>1</a></value></property>
                    <property type='f'><key><![CDATA[cornerRadius]]></key><value>1</value></property>
                    <property type='r'><key><![CDATA[frame]]></key><value><x>{rx}</x><y>{ry}</y><w>{rw}</w><h>{rh}</h></value></property>
                    <property type='b'><key><![CDATA[grabFocus]]></key><value>0</value></property>
                    <property type='b'><key><![CDATA[interactive]]></key><value>1</value></property>
                    <property type='b'><key><![CDATA[locked]]></key><value>0</value></property>
                    <property type='s'><key><![CDATA[name]]></key><value><![CDATA[pager1]]></value></property>
                    <property type='i'><key><![CDATA[orientation]]></key><value>0</value></property>
                    <property type='b'><key><![CDATA[outline]]></key><value>1</value></property>
                    <property type='i'><key><![CDATA[outlineStyle]]></key><value>0</value></property>
                    <property type='i'><key><![CDATA[pointerPriority]]></key><value>0</value></property>
                    <property type='i'><key><![CDATA[shape]]></key><value>1</value></property>
                    <property type='b'><key><![CDATA[tabLabels]]></key><value>1</value></property>
                    <property type='b'><key><![CDATA[tabbar]]></key><value>1</value></property>
                    <property type='b'><key><![CDATA[tabbarDoubleTap]]></key><value>0</value></property>
                    <property type='i'><key><![CDATA[tabbarSize]]></key><value>{TABBAR_SIZE}</value></property>
                    <property type='i'><key><![CDATA[textSizeOff]]></key><value>40</value></property>
                    <property type='i'><key><![CDATA[textSizeOn]]></key><value>40</value></property>
                    <property type='b'><key><![CDATA[visible]]></key><value>1</value></property>
                </properties>

                <values>
                    <value><key><![CDATA[page]]></key><locked>0</locked><lockedDefaultCurrent>0</lockedDefaultCurrent><default><![CDATA[0]]></default><defaultPull>0</defaultPull></value>
                    <value><key><![CDATA[touch]]></key><locked>0</locked><lockedDefaultCurrent>0</lockedDefaultCurrent><default><![CDATA[false]]></default><defaultPull>0</defaultPull></value>
                </values>

                <children>
{pages_xml}                </children>
            </node>
"""

def build_layout_xml_per_sheets(wb) -> str:
    root_id = new_id()
    rx, ry, rw, rh = ROOT_FRAME

    sheet_names = SHEETS if SHEETS is not None else wb.sheetnames
    pages_xml_list: List[str] = []
    page_index = 0

    for sname in sheet_names:
        if sname in SKIP_SHEETS:
            print(f"Überspringe Sheet: {sname}")
            continue

        ws = wb[sname]
        defs, tab_color = read_rows_from_ws(ws)

        page_index += 1
        inner = build_page_children_xml(defs)

        tab_label = sname

        pages_xml_list.append(
            page_group_xml(
                page_id=new_id(),
                page_name=ws.title,
                tab_label=tab_label,
                tab_color_hex=tab_color,
                inner_children=inner,
            )
        )

    pager = pager_xml(pager_id=new_id(), pages_xml="".join(pages_xml_list))

    return f"""<?xml version='1.0' encoding='UTF-8'?>
<lexml version='5'>
    <node ID='{root_id}' type='GROUP'>
        <properties>
            <property type='b'><key><![CDATA[background]]></key><value>1</value></property>
            <property type='c'><key><![CDATA[color]]></key><value><r>0</r><g>0</g><b>0</b><a>1</a></value></property>
            <property type='f'><key><![CDATA[cornerRadius]]></key><value>1</value></property>
            <property type='r'><key><![CDATA[frame]]></key><value><x>{rx}</x><y>{ry}</y><w>{rw}</w><h>{rh}</h></value></property>
            <property type='b'><key><![CDATA[grabFocus]]></key><value>0</value></property>
            <property type='b'><key><![CDATA[interactive]]></key><value>0</value></property>
            <property type='b'><key><![CDATA[locked]]></key><value>0</value></property>
            <property type='i'><key><![CDATA[orientation]]></key><value>0</value></property>
            <property type='b'><key><![CDATA[outline]]></key><value>1</value></property>
            <property type='i'><key><![CDATA[outlineStyle]]></key><value>0</value></property>
            <property type='i'><key><![CDATA[pointerPriority]]></key><value>0</value></property>
            <property type='i'><key><![CDATA[shape]]></key><value>1</value></property>
            <property type='b'><key><![CDATA[visible]]></key><value>1</value></property>
            <property type='s'><key><![CDATA[K1]]></key><value><![CDATA[{OUT_FILENAME}]]></value></property>
        </properties>

        <values>
            <value><key><![CDATA[touch]]></key><locked>0</locked><lockedDefaultCurrent>0</lockedDefaultCurrent><default><![CDATA[false]]></default><defaultPull>0</defaultPull></value>
        </values>

        <children>
{pager}        </children>
    </node>
</lexml>
"""

def choose_xlsx() -> Path:
    xlsx_files = sorted(BASE_DIR.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    if not xlsx_files:
        raise FileNotFoundError(f"Keine .xlsx Dateien in {BASE_DIR} gefunden.")

    if len(xlsx_files) == 1:
        print(f"Verwende: {xlsx_files[0].name}")
        return xlsx_files[0]

    print("Verfügbare Excel-Dateien:")
    for i, f in enumerate(xlsx_files, start=1):
        print(f"  {i}) {f.name}")

    while True:
        choice = input(f"Welche Datei verwenden? (1-{len(xlsx_files)}): ").strip()
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(xlsx_files):
                return xlsx_files[idx]
        except ValueError:
            pass
        print("Ungültige Eingabe, bitte nochmal.")


def main() -> None:
    xlsx_path = choose_xlsx()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    xml = build_layout_xml_per_sheets(wb)

    out_path = BASE_DIR / (xlsx_path.stem + ".tosc")
    out_path.write_text(xml, encoding="utf-8")
    print(f"Fertig: {out_path.resolve()}")

if __name__ == "__main__":
    main()
